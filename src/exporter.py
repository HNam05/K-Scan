"""
Finanzamt-tauglicher Report-Exporter fuer Mae Thai.
Erstellt:
  1. Excel-Monatsabrechnung (.xlsx) mit Belegnummer, Rahmen, SUM()-Formeln
  2. DIN A4 PDF-Bericht ueber ReportLab (kein GTK erforderlich)
"""
import os
import json
from datetime import datetime

# ─── Excel ────────────────────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ─── PDF (ReportLab) ──────────────────────────────────────────────────────────
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT


# ─── Helper ───────────────────────────────────────────────────────────────────

def _load_receipts(data_dir, month=None):
    """
    Laedt alle gespeicherten JSON-Belege, sortiert nach Datum.
    Optional: Filtert nach Monat im Format 'YYYY-MM'.
    """
    receipts = []
    for filename in os.listdir(data_dir):
        if not filename.endswith(".json"):
            continue
        path = os.path.join(data_dir, filename)
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
                data["_filename"] = filename
                receipts.append(data)
        except Exception:
            continue

    # Monatsfilter
    if month:
        try:
            filter_year, filter_month = int(month[:4]), int(month[5:7])
            filtered = []
            for r in receipts:
                datum_str = r.get("Datum", "")
                try:
                    parts = datum_str.split(".")
                    if len(parts) >= 3:
                        d_day, d_month, d_year = int(parts[0]), int(parts[1]), int(parts[2])
                        if d_year < 100:
                            d_year += 2000
                        if d_month == filter_month and d_year == filter_year:
                            filtered.append(r)
                except (ValueError, IndexError):
                    pass
            receipts = filtered
        except (ValueError, IndexError):
            pass  # Ungültiges Format → kein Filter

    # Sortierung nach Datum (älteste zuerst)
    def sort_key(r):
        try:
            parts = r.get("Datum", "01.01.2000").split(".")
            d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
            if y < 100:
                y += 2000
            return datetime(y, m, d)
        except Exception:
            return datetime(2000, 1, 1)

    receipts.sort(key=sort_key)
    return receipts


def _safe_float(value):
    """Konvertiert String/Float sicher in float."""
    try:
        return round(float(str(value).replace(",", ".")), 2)
    except (ValueError, TypeError):
        return 0.0


def _format_eur(value):
    """Formatiert float als Euro-String fuer PDF."""
    return f"{value:,.2f} €".replace(",", "X").replace(".", ",").replace("X", ".")


def _month_label(month_str=None):
    """Gibt 'März 2026' o.ä. zurueck."""
    MONTHS_DE = [
        "", "Januar", "Februar", "März", "April", "Mai", "Juni",
        "Juli", "August", "September", "Oktober", "November", "Dezember"
    ]
    if month_str:
        try:
            y, m = int(month_str[:4]), int(month_str[5:7])
            return f"{MONTHS_DE[m]} {y}"
        except Exception:
            pass
    now = datetime.now()
    return f"{MONTHS_DE[now.month]} {now.year}"


# ─── Excel ────────────────────────────────────────────────────────────────────

THIN = Side(style="thin", color="CCCCCC")
THICK = Side(style="medium", color="4F81BD")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_HEADER = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)

HEADER_FILL = PatternFill("solid", fgColor="1F3864")   # Dunkelblau
TOTAL_FILL  = PatternFill("solid", fgColor="D6E4F0")   # Hellblau
TITLE_FILL  = PatternFill("solid", fgColor="0C2461")   # Navy

MONEY_FMT = '#,##0.00 "€"'


def _xl_col(col):
    """Spaltenbuchstabe (1-basiert) → z.B. 1 → 'A'."""
    return get_column_letter(col)


class ReportExporter:
    def __init__(self):
        base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        self.data_dir   = os.path.join(base, "data")
        self.export_dir = os.path.join(base, "belege")
        os.makedirs(self.export_dir, exist_ok=True)

    # ── Excel ──────────────────────────────────────────────────────────────

    def generate_excel_report(self, month=None):
        """
        Erstellt eine professionelle Excel-Monatsabrechnung.
        :param month: z.B. '2026-03' oder None fuer aktuellen Monat
        :return: Pfad zur erstellten .xlsx-Datei
        """
        if month is None:
            month = datetime.now().strftime("%Y-%m")

        receipts = _load_receipts(self.data_dir, month)
        if not receipts:
            raise ValueError(
                f"Keine Belege fuer {_month_label(month)} vorhanden."
            )

        wb = Workbook()
        ws = wb.active
        ws.title = "Monatsabrechnung"

        month_label = _month_label(month)
        now_str = datetime.now().strftime("%d.%m.%Y %H:%M")

        # ── Zeile 1: Titelbereich ──────────────────────────────────────────
        ws.merge_cells("A1:I1")
        title_cell = ws["A1"]
        title_cell.value = f"Mae Thai — Monatsabrechnung {month_label}"
        title_cell.font = Font(bold=True, size=14, color="FFFFFF",
                               name="Calibri")
        title_cell.fill = TITLE_FILL
        title_cell.alignment = Alignment(horizontal="center",
                                         vertical="center")
        ws.row_dimensions[1].height = 36

        # ── Zeile 2: Erstellt am ──────────────────────────────────────────
        ws.merge_cells("A2:I2")
        sub_cell = ws["A2"]
        sub_cell.value = f"Betriebsstätte: Mae Thai Restaurant  •  Erstellt am: {now_str}"
        sub_cell.font = Font(italic=True, size=9, color="666666", name="Calibri")
        sub_cell.fill = PatternFill("solid", fgColor="EBF2FF")
        sub_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 18

        # ── Zeile 3: Leerzeile ────────────────────────────────────────────
        ws.row_dimensions[3].height = 8

        # ── Zeile 4: Kopfzeile ────────────────────────────────────────────
        HEADERS = [
            "Beleg-Nr.",  # A
            "Datum",      # B
            "Händler",    # C
            "Kategorie",  # D
            "Brutto (€)", # E
            "Netto (€)",  # F
            "MwSt 7% (€)",# G
            "MwSt 19%(€)",# H
            "Bemerkung",  # I
        ]
        header_row = 4
        for col_idx, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            cell.border = BORDER_THIN
        ws.row_dimensions[header_row].height = 30

        # ── Spaltenbreiten ────────────────────────────────────────────────
        col_widths = [11, 12, 22, 16, 14, 14, 14, 14, 20]
        for idx, w in enumerate(col_widths, start=1):
            ws.column_dimensions[_xl_col(idx)].width = w

        # ── Datenzeilen ───────────────────────────────────────────────────
        totals = {"Brutto": 0.0, "Netto": 0.0, "Steuer_7": 0.0, "Steuer_19": 0.0}
        data_start_row = header_row + 1

        for r_idx, receipt in enumerate(receipts, start=data_start_row):
            beleg_nr = str(r_idx - data_start_row + 1).zfill(3)
            haendler = (receipt.get("Haendler")
                        or receipt.get("Händler", "Unbekannt"))
            datum    = receipt.get("Datum", "")
            kategorie = receipt.get("Kategorie", "Lebensmittel")

            brutto    = _safe_float(receipt.get("Brutto",    0))
            netto     = _safe_float(receipt.get("Netto",     0))
            steuer_7  = _safe_float(receipt.get("Steuer_7",  0))
            steuer_19 = _safe_float(receipt.get("Steuer_19", 0))

            row_vals = [beleg_nr, datum, haendler, kategorie,
                        brutto, netto, steuer_7, steuer_19, ""]

            for c_idx, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = val
                cell.border = BORDER_THIN
                cell.font = Font(size=9, name="Calibri")

                if c_idx in (5, 6, 7, 8):          # Geldspalten
                    cell.number_format = MONEY_FMT
                    cell.alignment = Alignment(horizontal="right",
                                               vertical="center")
                elif c_idx in (1, 2):              # Nr. + Datum
                    cell.alignment = Alignment(horizontal="center",
                                               vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left",
                                               vertical="center",
                                               wrap_text=(c_idx == 9))

            # Zebra-Streifen
            if (r_idx - data_start_row) % 2 == 1:
                row_fill = PatternFill("solid", fgColor="F5F8FF")
                for c_idx in range(1, 10):
                    ws.cell(row=r_idx, column=c_idx).fill = row_fill

            ws.row_dimensions[r_idx].height = 18

            totals["Brutto"]    += brutto
            totals["Netto"]     += netto
            totals["Steuer_7"]  += steuer_7
            totals["Steuer_19"] += steuer_19

        # ── Summenzeile ───────────────────────────────────────────────────
        sum_row = data_start_row + len(receipts)
        ws.merge_cells(f"A{sum_row}:D{sum_row}")
        label_cell = ws.cell(row=sum_row, column=1)
        label_cell.value = f"GESAMTSUMME — {month_label} ({len(receipts)} Belege)"
        label_cell.font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        label_cell.fill = HEADER_FILL
        label_cell.alignment = Alignment(horizontal="left", vertical="center")
        label_cell.border = BORDER_THIN

        for merge_col in (2, 3, 4):
            ws.cell(row=sum_row, column=merge_col).fill = HEADER_FILL

        # SUM()-Formeln für die vier Geldspalten (E, F, G, H)
        for col_idx, key in zip(range(5, 9), ["Brutto", "Netto", "Steuer_7", "Steuer_19"]):
            col_letter = _xl_col(col_idx)
            cell = ws.cell(row=sum_row, column=col_idx)
            cell.value = (f"=SUM({col_letter}{data_start_row}"
                          f":{col_letter}{sum_row - 1})")
            cell.number_format = MONEY_FMT
            cell.font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = BORDER_THIN

        ws.cell(row=sum_row, column=9).fill = HEADER_FILL
        ws.cell(row=sum_row, column=9).border = BORDER_THIN
        ws.row_dimensions[sum_row].height = 22

        # ── Fenster einfrieren ab Zeile 5 ─────────────────────────────────
        ws.freeze_panes = f"A{data_start_row}"

        # ── Druckbereich setzen (DIN A4) ──────────────────────────────────
        ws.print_area = f"A1:I{sum_row}"
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_setup.paperSize  = ws.PAPERSIZE_A4
        ws.page_setup.orientation = "landscape"
        ws.print_title_rows = f"1:{header_row}"

        # ── Speichern ─────────────────────────────────────────────────────
        filename = f"Mae_Thai_Abrechnung_{month}.xlsx"
        output_path = os.path.join(self.export_dir, filename)
        wb.save(output_path)
        return output_path

    # ── PDF ────────────────────────────────────────────────────────────────

    def generate_pdf_report(self, month=None):
        """
        Erstellt eine DIN A4 PDF-Monatsabrechnung (ReportLab).
        :param month: z.B. '2026-03' oder None fuer aktuellen Monat
        :return: Pfad zur .pdf-Datei
        """
        if month is None:
            month = datetime.now().strftime("%Y-%m")

        receipts = _load_receipts(self.data_dir, month)
        if not receipts:
            raise ValueError(
                f"Keine Belege fuer {_month_label(month)} vorhanden."
            )

        month_label = _month_label(month)
        now_str = datetime.now().strftime("%d.%m.%Y %H:%M")

        filename = f"Mae_Thai_Abrechnung_{month}.pdf"
        output_path = os.path.join(self.export_dir, filename)

        doc = SimpleDocTemplate(
            output_path,
            pagesize=A4,
            rightMargin=1.5 * cm,
            leftMargin=1.5 * cm,
            topMargin=2 * cm,
            bottomMargin=2 * cm,
            title=f"Mae Thai — Monatsabrechnung {month_label}",
            author="K-Scan Intelligence",
        )

        styles = getSampleStyleSheet()

        style_title = ParagraphStyle(
            "MaeThaiTitle",
            parent=styles["Heading1"],
            fontSize=16,
            textColor=colors.HexColor("#1F3864"),
            spaceAfter=4,
            alignment=TA_CENTER,
            fontName="Helvetica-Bold",
        )
        style_sub = ParagraphStyle(
            "MaeThaiSub",
            parent=styles["Normal"],
            fontSize=8,
            textColor=colors.grey,
            spaceAfter=12,
            alignment=TA_CENTER,
        )

        story = []

        # ── Header ────────────────────────────────────────────────────────
        story.append(Paragraph(f"Mae Thai — Monatsabrechnung {month_label}", style_title))
        story.append(Paragraph(
            f"Betriebsstätte: Mae Thai Restaurant &nbsp;&nbsp;|&nbsp;&nbsp; "
            f"Erstellt am: {now_str}", style_sub
        ))
        story.append(HRFlowable(width="100%", thickness=1.5,
                                 color=colors.HexColor("#1F3864"),
                                 spaceAfter=12))

        # ── Tabellen-Header ───────────────────────────────────────────────
        col_headers = [
            "Nr.", "Datum", "Händler", "Kategorie",
            "Brutto", "Netto", "MwSt 7%", "MwSt 19%"
        ]

        table_data = [col_headers]

        totals = {"brutto": 0.0, "netto": 0.0, "s7": 0.0, "s19": 0.0}

        for idx, r in enumerate(receipts, start=1):
            haendler  = r.get("Haendler") or r.get("Händler", "Unbekannt")
            datum     = r.get("Datum", "")
            kategorie = r.get("Kategorie", "Lebensmittel")
            brutto    = _safe_float(r.get("Brutto",    0))
            netto     = _safe_float(r.get("Netto",     0))
            s7        = _safe_float(r.get("Steuer_7",  0))
            s19       = _safe_float(r.get("Steuer_19", 0))

            table_data.append([
                str(idx).zfill(3),
                datum,
                haendler[:30],        # Lange Namen abschneiden
                kategorie,
                _format_eur(brutto),
                _format_eur(netto),
                _format_eur(s7),
                _format_eur(s19),
            ])

            totals["brutto"] += brutto
            totals["netto"]  += netto
            totals["s7"]     += s7
            totals["s19"]    += s19

        # ── Summenzeile ───────────────────────────────────────────────────
        table_data.append([
            "", f"{len(receipts)} Belege",
            "GESAMT", "",
            _format_eur(totals["brutto"]),
            _format_eur(totals["netto"]),
            _format_eur(totals["s7"]),
            _format_eur(totals["s19"]),
        ])

        # ── Spaltenbreiten (A4 = 17 cm nutzbar im Hochformat) ─────────────
        page_w = A4[0] - 3 * cm   # nutzbare Breite
        col_widths_pdf = [
            page_w * 0.05,   # Nr.
            page_w * 0.10,   # Datum
            page_w * 0.22,   # Händler
            page_w * 0.13,   # Kategorie
            page_w * 0.125,  # Brutto
            page_w * 0.125,  # Netto
            page_w * 0.125,  # MwSt 7%
            page_w * 0.125,  # MwSt 19%
        ]

        table = Table(table_data, colWidths=col_widths_pdf, repeatRows=1)

        navy    = colors.HexColor("#1F3864")
        lt_blue = colors.HexColor("#D6E4F0")
        stripe  = colors.HexColor("#F5F8FF")
        last    = len(table_data) - 1   # Index der Summenzeile

        ts = TableStyle([
            # Header
            ("BACKGROUND",    (0, 0), (-1, 0),    navy),
            ("TEXTCOLOR",     (0, 0), (-1, 0),    colors.white),
            ("FONTNAME",      (0, 0), (-1, 0),    "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, 0),    8),
            ("ALIGN",         (0, 0), (-1, 0),    "CENTER"),
            ("VALIGN",        (0, 0), (-1, 0),    "MIDDLE"),
            ("ROWBACKGROUND", (0, 0), (-1, 0),    navy),
            ("TOPPADDING",    (0, 0), (-1, 0),    6),
            ("BOTTOMPADDING", (0, 0), (-1, 0),    6),

            # Daten-Zeilen
            ("FONTNAME",  (0, 1), (-1, last - 1), "Helvetica"),
            ("FONTSIZE",  (0, 1), (-1, last - 1), 8),
            ("VALIGN",    (0, 1), (-1, last - 1), "MIDDLE"),
            ("TOPPADDING",    (0, 1), (-1, last - 1), 4),
            ("BOTTOMPADDING", (0, 1), (-1, last - 1), 4),

            # Zebra-Streifen (gerade Zeilen, 0-basiert)
            *[("BACKGROUND", (0, i), (-1, i), stripe)
              for i in range(2, last, 2)],

            # Geldspalten rechts ausrichten
            ("ALIGN",  (4, 1), (-1, last - 1), "RIGHT"),
            ("ALIGN",  (0, 1), (0, last - 1),  "CENTER"),  # Nr.
            ("ALIGN",  (1, 1), (1, last - 1),  "CENTER"),  # Datum

            # Summenzeile
            ("BACKGROUND", (0, last), (-1, last), lt_blue),
            ("FONTNAME",   (0, last), (-1, last), "Helvetica-Bold"),
            ("FONTSIZE",   (0, last), (-1, last), 8.5),
            ("ALIGN",      (4, last), (-1, last), "RIGHT"),
            ("LINEABOVE",  (0, last), (-1, last), 1.5, navy),

            # äußerer Rahmen + innere Linien
            ("BOX",       (0, 0), (-1, -1), 1,   navy),
            ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CCCCCC")),
        ])
        table.setStyle(ts)
        story.append(table)

        # ── Hinweis unter der Tabelle ─────────────────────────────────────
        story.append(Spacer(1, 0.6 * cm))
        story.append(HRFlowable(width="100%", thickness=0.5,
                                 color=colors.lightgrey))
        story.append(Spacer(1, 0.3 * cm))
        hint_style = ParagraphStyle(
            "Hint", parent=styles["Normal"],
            fontSize=7, textColor=colors.grey, leading=10
        )
        story.append(Paragraph(
            "Alle Beträge basieren auf automatisch erkannten Kassenbons und "
            "wurden manuell geprüft. Diese Aufstellung dient als Grundlage "
            "für die Betriebsausgaben-Erfassung gemäß §4 EStG.",
            hint_style
        ))

        # ── Footer-Funktion ───────────────────────────────────────────────
        def _add_footer(canvas, doc):
            canvas.saveState()
            canvas.setFont("Helvetica", 7)
            canvas.setFillColor(colors.grey)
            page_num = canvas.getPageNumber()
            w, _ = A4
            canvas.drawString(1.5 * cm, 1.2 * cm,
                              f"Mae Thai — Monatsabrechnung {month_label}")
            canvas.drawRightString(w - 1.5 * cm, 1.2 * cm,
                                   f"Seite {page_num}")
            canvas.restoreState()

        doc.build(story, onFirstPage=_add_footer, onLaterPages=_add_footer)
        return output_path


# ── CLI-Test ──────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    exp = ReportExporter()
    try:
        xl = exp.generate_excel_report()
        print(f"✅ Excel: {xl}")
    except Exception as e:
        print(f"❌ Excel Fehler: {e}")

    try:
        pdf = exp.generate_pdf_report()
        print(f"✅ PDF:   {pdf}")
    except Exception as e:
        print(f"❌ PDF Fehler: {e}")
